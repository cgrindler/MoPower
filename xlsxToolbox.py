import openpyxl as pyxl
import os
import PowerToolbox as ptbx
import pandas as pd 
import os  
import cmath

class createWorkBook():
    
    def __init__(self,sys,fmuName,Modus,dynamic):
        
        currDir = os.path.dirname(os.path.realpath(__file__))
        self.Modus = Modus
        self.filePath = "Config/"+ fmuName + "/"
        self.sys = sys
        self.base = self.sys["baseMVA"]
        self.dynamic = dynamic
        self.metaDict = dict()
        self.solverDict = dict()

        self.metaDict["Object"] = fmuName
        self.metaDict["ModelName"] = fmuName
        self.metaDict["Host"] = "localhost:6379"
        self.metaDict["SolverVersion"] = "1.9.8"
        self.metaDict["SolverName"] = "HQP"
        self.metaDict["SolverIteration"] = 50
        self.metaDict["SolverIterationLimit"] = 100
        
        #wrkPath = "C:\Program Files (x86)\ABB Industrial IT\Optimize IT\Dynamic Optimization\work"
        wrkPath = os.getcwd() + "/Config"
        if self.dynamic:
            tmpName = "Dynamic Optimization.xlsx"
        else: tmpName = "Static Optimization.xlsx"

        self.wbName = fmuName + "_" + tmpName 
                      
        if os.path.isdir(wrkPath):            
            self.savePath = wrkPath + "/" + self.wbName
            if os.path.isfile(self.savePath):
                self.wb = pd.read_excel(self.savePath,sheetname=["meta","input","output","Solver"],header=None)
            else:  
                xlwriter = pd.ExcelWriter(self.savePath, engine='xlsxwriter')
                self.wb = xlwriter.book #new empty worbook -toDo

        elif os.path.isfile(currDir + "/"+ self.wbName): 
            tmpPath = currDir + "/" + self.wbName
            self.wb = pd.read_excel(tmpPath,sheetname=["meta","input","output","Solver"],header=None)
            self.savePath = currDir + "/Config/" + fmuName + "/"+ self.wbName
        else: print("\nKeine DynamicOptimizer Konfiguration gefunden. Bitte laden Sie die gewünschte Konfiguration in den aktuellen Projektordner.\n") 

    def setDependencies(self):
                
        metaDeps = self.wb["meta"].values
        inputDeps = self.wb["input"].values
        outputDeps = self.wb["output"].values
        solverDeps = self.wb["Solver"].values
        self.setDicts()
        self.writeKeys(metaDeps,self.metaDict,"meta")
        self.writeKeys(inputDeps,self.inDict,"input")
        self.writeKeys(outputDeps,self.outDict,"output")

    def setDicts(self):

        self.outDict = dict()
        self.inDict = dict()

        #Power Limits
        for i, elements in enumerate(self.sys["gen"]):            
            geni = ptbx.convertGenElements(True,elements)
            
            No = str(geni["No"])
            Qmin = geni["Qmin"]/self.base
            Qmax = geni["Qmax"]/self.base
            Pmin = geni["Pmin"]/self.base
            Pmax = geni["Pmax"]/self.base
            
            if self.Modus == 5:
                self.inDict["Pg"+No] = []
                self.inDict["Pg"+No].append([Pmin, Pmax])

            else:
                self.outDict["Sg"+No+"[1]"] = []
                self.outDict["Sg"+No+"[2]"] = []
                self.outDict["Sg"+No+"[1]"].append([Pmin, Pmax])
                self.outDict["Sg"+No+"[2]"].append([Qmin, Qmax])

            lenCosts = len(self.sys["gencost"][0])
            if lenCosts != 7 or self.sys["gencost"][i][lenCosts-1] !=0 or self.sys["gencost"][i][0] != 2:
                print("\n\n => HQP isn't supporting the defined cost structure - Please check your PTI!")
            
            
            c1 =  self.sys["gencost"][i][lenCosts-2]*self.base
            c2 =  self.sys["gencost"][i][lenCosts-3]*self.base**2
            if self.Modus == 5:
                self.inDict["Pg"+No].append([c1, c2])
                self.inDict["Pg"+No].append(1/self.base)
            else:             
                self.outDict["Sg"+No+"[1]"].append([c1, c2])
        
        # Voltage Limits
        for elements in self.sys["bus"]:
            
            busi = ptbx.convertBusElements(True,elements)
            No = str(busi["No"])
            Vmax = busi["Vmax"]
            Vmin = busi["Vmin"]
            type = busi["busType"]
            Vabs = busi["Vabs"]
            phi = busi["phi"]
            Qload = busi["Qload"]/self.base
            Pload = busi["Pload"]/self.base
            Vr = cmath.rect(float(Vabs),float(phi))  

            if self.Modus == 2:
                if not type == 3:
                    self.outDict["cns_Vabs"+No] = [Vmin**2, Vmax**2]
                  
                if type == 2:
                    self.inDict["Vr"+No+"[1]"] = [1, Vr.real]
                    self.inDict["Vr"+No+"[2]"] = [1, Vr.imag]  
                if type == 1: 
                    if Pload == 0 and Qload == 0:
                        self.outDict["Vr"+No+"[1]"] = 0
                        self.outDict["Vr"+No+"[2]"] = 0
                    else:
                        self.inDict["Vr"+No+"[1]"] = [1, Vr.real]
                        self.inDict["Vr"+No+"[2]"] = [1, Vr.imag]
                if type == 3:
                    self.inDict["Vr"+No+"[1]"] = [0, Vr.real]
                    self.inDict["Vr"+No+"[2]"] = [0, Vr.imag]  

            elif self.Modus == 5:
                if type == 3:
                    self.outDict["Phi" + No] = 0
            else:
                if type == 2:
                    self.inDict["Vp"+No+"[1]"] = [Vmin**2, Vmax**2, 1]

                if type == 1:
                    self.outDict["Vp"+No+"[1]"] = [Vmin**2, Vmax**2]

            
            if Pload != 0 or Qload != 0:
                if self.Modus == 5:
                    self.inDict["Pl"+No] = Pload
                else:
                    self.inDict["Sl"+No+"[1]"] = [0, Pload]
                    self.inDict["Sl"+No+"[2]"] = [0, Qload]
                    if type == 1:
                        self.outDict["ceq_Sl"+No+"[1]"] = [0, 0]
                        self.outDict["ceq_Sl"+No+"[2]"] = [0, 0]

        #Flow Limits:
        for i, elements in enumerate(self.sys["branch"]):
            
            branchi = ptbx.convertBranchElements(True,elements)
            fr = branchi["from"]
            to = branchi["to"]
            rateA = branchi["rateA"]
            if rateA != 0:
                if not self.Modus == 5:
                    self.outDict["cns_I"+str(fr)+"_"+str(to)] = ["-Inf", (rateA/int(self.base))**2]

    def writeKeys(self,deps,xlDict,name):

        headers = deps[0]
        hDict=self.getColumns(headers)
        for i, elements in enumerate(deps):                    
            if name == "meta":

                metaKey = elements[0]
                if metaKey in xlDict:
                    self.wb[name].values[i][2] = xlDict[metaKey]

            if name == "input":
                inputKey = elements[2]
                if inputKey in self.inDict:
                    if self.Modus == 2:
                        self.wb[name].values[i][hDict["u_active"]] = xlDict[inputKey][0]
                        self.wb[name].values[i][hDict["u_start"]] = xlDict[inputKey][1]  
                    elif self.Modus == 1:
                        if inputKey[0] == "V":
                            self.wb[name].values[i][hDict["u_active"]] = xlDict[inputKey][2] # 1: SOH, 2: u_connection 3: varname 4:u 5:u_start 6:u_order 7:u_active 8: u0_nfixed 9: u_nominal 10: u_min 11: u_max
                            self.wb[name].values[i][hDict["u_max"]] = xlDict[inputKey][1] 
                            self.wb[name].values[i][hDict["u_min"]] = xlDict[inputKey][0]
                        else:
                            self.wb[name].values[i][hDict["u_active"]] = xlDict[inputKey]

                    elif self.Modus == 5:
                            if inputKey[:2] == "Pg":
                                self.wb[name].values[i][hDict["u_active"]] = 1
                                self.wb[name].values[i][hDict["u_nominal"]] = xlDict[inputKey][2]
                                self.wb[name].values[i][hDict["u_max"]] = xlDict[inputKey][0][1] 
                                self.wb[name].values[i][hDict["u_min"]] = xlDict[inputKey][0][0]
                                self.wb[name].values[i][hDict["u_weight1"]] = xlDict[inputKey][1][0] 
                                self.wb[name].values[i][hDict["u_weight2"]] = xlDict[inputKey][1][1]
                            else:
                                self.wb[name].values[i][hDict["u_active"]] = 0

            if name == "output":
                
                outputKey = elements[2]
                if self.Modus == 5:
                    if outputKey in self.outDict:
                        self.wb[name].values[i][hDict["y_min"]] = 0
                        self.wb[name].values[i][hDict["y_max"]] = 0
                else:
                    if outputKey in self.outDict:

                        if not outputKey[0] == "V" and not outputKey[0] == "S" and not outputKey[:3] == "ceq":                    
                            self.wb[name].values[i][hDict["y_min"]] = float(self.outDict[outputKey][0])
                            self.wb[name].values[i][hDict["y_max"]] = float(self.outDict[outputKey][1])

                        elif outputKey[0] == "S":
                            if outputKey[-2] == "1":
                                self.wb[name].values[i][hDict["y_min"]] = float(self.outDict[outputKey][0][0]) 
                                self.wb[name].values[i][hDict["y_max"]] = float(self.outDict[outputKey][0][1]) 
                                self.wb[name].values[i][hDict["y0_weight1"]] = self.outDict[outputKey][1][0]*100 
                                self.wb[name].values[i][hDict["y0_weight2"]] = self.outDict[outputKey][1][1]*10000 
                                #self.wb[name].values[i][hDict["y0_weight1"]] = self.outDict[outputKey][1][0] #
                                #self.wb[name].values[i][hDict["y0_weight2"]] = self.outDict[outputKey][1][1]
                                #self.wb[name].values[i][hDict["y_nominal"]] = float(1/self.base) 
                            else:  
                                self.wb[name].values[i][hDict["y_min"]] = float(self.outDict[outputKey][0][0]) 
                                self.wb[name].values[i][hDict["y_max"]] = float(self.outDict[outputKey][0][1]) 

                        elif outputKey[:3] == "ceq":

                                #self.wb[name].values[i][hDict["y_min"]] = 0 #softmin 
                                #self.wb[name].values[i][hDict["y_max"]] = 0 #softmax 
                                self.wb[name].values[i][hDict["y_soft_min"]] = 0 #softmin 
                                self.wb[name].values[i][hDict["y_soft_max"]] = 0 #softmax 
                                self.wb[name].values[i][hDict["y_soft_weight1"]] = 1000 #softweight1
                                self.wb[name].values[i][hDict["y_soft_weight2"]] = 100  #softweight2
                                                                                                                                                   
    def saveWorkbook(self):
        
        if not os.path.isdir(self.filePath):
            os.mkdir(self.filePath)      
            
        book = pyxl.load_workbook(self.savePath)
        writer = pd.ExcelWriter(self.savePath, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
                
        self.wb["meta"].to_excel(writer, sheet_name ="meta",merge_cells = False,index = False,header = False)
        self.wb["input"].to_excel(writer, sheet_name ="input",merge_cells = False,index = False,header = False) 
        self.wb["output"].to_excel(writer, sheet_name ="output",merge_cells = False,index = False,header = False) 
        self.wb["Solver"].to_excel(writer, sheet_name ="Solver",merge_cells = False,index = False,header = False)  
        writer.save()

    def getColumns(self,headers):
        hDict = dict()
        for i,elements in enumerate(headers):
            hDict[elements] = i
        return hDict



